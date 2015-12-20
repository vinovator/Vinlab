# Python 2.7.6
# extractXmlFromXl.py

"""
Script to download all xml files from the urls specified in the excel report
Requires requests and openpyxl modules
"""

import requests
from requests.auth import HTTPDigestAuth # To access url with digest authentication
import os
import openpyxl as op # To iterate through the excel report
import logging
import Tkinter as tk # Expose GUI to user to grab report details
import tkFileDialog as tkf # To open file location
import tkMessageBox as tkm # To show warning/ error messages


# Basic logging config
LOG_FILE_NAME = "xml_download.log"
logging.basicConfig(filename = LOG_FILE_NAME, 
					format = "%(asctime)s: %(levelname)s: %(funcName)s: %(lineno)d: %(message)s",
					datefmt = "%m-%d-%Y %I:%M:%S %p",
					filemode = "w", 
					level = logging.WARNING) # All DEBUG logs will be skipped

					
class xmlExtractor():

	def __init__(self):
		"""
		Initialize all common parameters
		"""
		
		self.root = tk.Tk()
		self.root.title("Downlod XML files")
		self.root.minsize(width=400, height=250)
		self.REPORT_FILE_NAME = None
		self.REPORT_SHEET_NAME = None
		self.XML_FOLDER = None
		
	
	def upload_file(self):
		"""
		Upload button click event handler
		"""
		
		uploaded_file = tkf.askopenfilename() # this is the uploaded file
		
		if not (uploaded_file):
			tkm.showerror(title="Error", message="Please select a file to upload")
		elif((uploaded_file.split(".")[-1]) not in ("xlsx", "xls")):
			tkm.showerror(title="Error", message="Please upload a valid file of xlsx format")
		else:
			self.REPORT_FILE_NAME = uploaded_file # Name of the report with full path
			self.REPORT_SHEET_NAME = "Metadata Extraction" # Name of the worksheet with report data
			self.destfile_status_lbl.config(text="{} loaded".format(uploaded_file.split("/")[-1]))

		
	def select_folder(self):
		"""
		Select folder button click event handler
		"""
		
		folder_name = tkf.askdirectory() # Get the destination folder name
		
		if (folder_name):
			self.XML_FOLDER = folder_name
			self.folder_name_status.config(text="XML files will be saved in the folder {}".format(self.XML_FOLDER))
			if not os.path.exists(self.XML_FOLDER):
				os.makedirs(self.XML_FOLDER)
		else:
			tkm.showerror(title="Error", message="Please select a destiation Folder")

			
	def submit_click(self):
		"""
		Submit button click event handler	
		"""
		# Get credentials from user
		user = self.username_entry.get()
		pwd = self.password_entry.get()

		# At press of submit check if all mandator fields are provided
		if (user and pwd and self.XML_FOLDER and self.REPORT_FILE_NAME):
			self.username = user
			self.password = pwd
			# Close window once processing complete				
			self.root.destroy()
		else:
			tkm.showerror(title="Error", message="Mandatory fields cannot be empty")	

			
	def get_report_details_gui(self):
		"""
		Method Exposing GUI to user
		"""
		
		username_lbl = tk.Label(self.root, text="Username", padx=2, pady=5)
		self.username_entry = tk.Entry(self.root)
		
		password_lbl = tk.Label(self.root, text="Password", padx=2, pady=5)
		self.password_entry = tk.Entry(self.root, show="*")
		
		upload_lbl = tk.Label(self.root, text="Upload the report", padx=2, pady=5)
		
		open_file_button = tk.Button(self.root, text="Upload File", command=self.upload_file, width=17, pady=5)
		
		self.destfile_status_lbl = tk.Label(self.root, wraplength=400, fg="blue", pady=5)
		
		folder_name_lbl = tk.Label(self.root, text="Select folder to download xml files", padx=2, pady=5)
		self.folder_name_button = tk.Button(self.root, text="Select Folder", command=self.select_folder, width=17, pady=5)
		self.folder_name_status = tk.Label(self.root, wraplength=400, fg="blue", pady=5)
		
		submit_button = tk.Button(self.root, text="Submit", command=self.submit_click, width=20, pady=5)
		
		username_lbl.grid(row=0, column=0, sticky="E")
		self.username_entry.grid(row=0, column=1)
		password_lbl.grid(row=1, column=0, sticky="E")
		self.password_entry.grid(row=1, column=1)
		upload_lbl.grid(row=2, column=0, sticky="E")
		open_file_button.grid(row=2, column=1, columnspan=2)
		self.destfile_status_lbl.grid(row=3, column=0, columnspan=2)
		folder_name_lbl.grid(row=4, column=0, sticky="E")
		self.folder_name_button.grid(row=4, column=1, columnspan=2)
		self.folder_name_status.grid(row=5, column=0, columnspan=2, rowspan=2)
		submit_button.grid(row=8, column=0, columnspan=2)
		
		self.root.mainloop()

		
	def download_xml(self, url, output_file_name):
		try:
			response = requests.get(url, auth=HTTPDigestAuth(self.username, self.password), verify=True)
		except Exception as e:
			logging.exception("Error while accessing the url {}".format(url))
			print("Error while accessing the url {}".format(url))

		try:
			with open(os.path.join(self.XML_FOLDER + "/" + output_file_name), "wb") as xml_ouput:
				for chunk in response.iter_content(chunk_size=1024):
					if chunk:
						xml_ouput.write(chunk)
			print ("{0} saved in folder {1}".format(output_file_name, self.XML_FOLDER))
		except Exception as e:
			logging.exception("Error dowloading content from {}".format(url))
			print("Error dowloading content from {}".format(url))

		
	def iterate_report(self):
		wb = op.load_workbook(self.REPORT_FILE_NAME)
		sheet = wb.get_sheet_by_name(self.REPORT_SHEET_NAME)
		
		max_row = sheet.get_highest_row()
		max_column = sheet.get_highest_column()
		
		print("max row: {0}, max column: {1}".format(max_row, max_column))
		
		for row in range(2, max_row + 1):
			if (sheet.cell(row = row, column = 8).value == "FinishedSuccessfully"):
				item = sheet.cell(row = row, column = 1).value.split(":")[1]
				supplier = sheet.cell(row = row, column = 4).value.split("/")[0]
				asset = sheet.cell(row = row, column = 10).value
				url = sheet.cell(row = row, column = 11).value
				
				try:
					self.download_xml(url, item + "_" + supplier + "_" + asset + ".xml")
				except Exception as e:
					logging.exception("Error extracting xml for pii:{0} from {1} for asset type {2}".format(item, supplier, asset))
					continue

			
if __name__ == "__main__":

	xe = xmlExtractor()
	
	xe.get_report_details_gui()
	
	xe.iterate_report()
	
	print("Done. Refer to {} to check the job status".format(LOG_FILE_NAME))