#python 3
#Read the list of names and email ids from an excel file and populate them in a text file

import openpyxl
import os
import time

def fetchUnpaidMembers(max_row, max_col, sheet):
	member_count = 0
	mydict = {}
	for r in range(2, max_row + 1):
		if sheet.cell(row = r, column = max_col).value is None:
			member_count += 1
			name = sheet.cell(row = r, column = 1).value
			email = sheet.cell(row = r, column = 2).value
			mydict[name] = email
	print ("There are {0} unpaid members".format(member_count))
	return mydict
	
def logUnpaidMembers(logdict, path):
	#Dictionary will return false if empty
	if logdict:
		print ("Logging begins")
		#log the data in a text file
		with open(path + "\\" + "UnpaidMembers.log", "a", encoding = "utf-8") as mylog:
			mylog.write("***** Begin log on " + time.ctime() + " *****")
			mylog.write("\n")
			for name, email in logdict.items():
				mylog.write(name + " : " + email)
				mylog.write("\n")
			mylog.write("***** End *****")
			mylog.write("\n\n")
		print ("Logging ends")
	else:
		print("No unpaid members to log")
	

def main():

	excel_path = ".\EXCEL"
	excel_file = "duedata.xlsx"

	wb = openpyxl.load_workbook(excel_path + "\\" + excel_file)
	sheet = wb.get_sheet_by_name("Sheet1")

	last_member = sheet.get_highest_row()
	latest_month = sheet.get_highest_column()

	#print(last_member, latest_month)

	#This empty dict will carry the name and email of those who have not paid dues for the latest month
	due_dict = {}

	#Fetch the list of unpaid members by checking the latest month
	due_dict = fetchUnpaidMembers(last_member, latest_month, sheet)

	'''for name, email in due_dict.items():
		print (name, email)'''
		
	#TODO: Extend the program by sending an email and sms to the unpaid members
		
	#Log the details of the unpaid members in a text file
	logUnpaidMembers(due_dict, excel_path)
	

if __name__ == "__main__":
	main()