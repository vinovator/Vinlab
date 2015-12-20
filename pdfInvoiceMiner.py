__author__ = 'Vinoth_Subramanian'
# Python3
# pdfInvoiceMiner.py

# Program to extract the client info and invoice no from a bunch of invoice pdf files
# pdfminer3k library is used to extract text from pdf
# PyPDF2 library does not extract the text from pdf properly
# place all the invoice pdf files within a folder named "INVOICE"
# place an excel file named "invoice_info.xlsx" in the parent folder of "INVOICE"
# First column - invoice no; Second column - client details

import PyPDF2
from pdfminer.pdfparser import PDFParser, PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import PDFPageAggregator
from pdfminer.layout import LAParams, LTTextBox, LTTextLine
import os
import openpyxl


# The folder path where are all the invoices are placed
invoice_path = ".\COMPANY_NAME\INVOICE" # Company name masked - replace (parent folder name)
# The excel file where the invoice information is extracted
invoice_info_xl = ".\COMPANY_NAME\invoice_info.xlsx" # Company name masked - replace (parent folder name)

# Within the pdf, the start block and end block within which the client information is placed
client_start = "Ph: 9999999999, 9999999999" # Phone number masked. Replace with actual phone number in invoice
client_end = "Qty Description"

# From the pdf file name, the start block and end block within which the invoice no is placed
invoice_start = "Invoice-"
invoice_end = ".pdf"


# Given a string, start block and end block; find the substrinng between them
def extract_info(strInfo, strStart, strEnd):
    return strInfo[strInfo.find(strStart) + len(strStart) : strInfo.rfind(strEnd)]


# Open the excel file and update the invoice number in column 1 and client info in column 2
def write_excel(client, invoice_no):
    wb = openpyxl.load_workbook(invoice_info_xl)
    sheet = wb.get_sheet_by_name("Sheet1")
    last_entry = sheet.get_highest_row()
    sheet.cell(row = last_entry+1, column=1).value = int(invoice_no)
    sheet.cell(row = last_entry+1, column=2).value = client
    wb.save(invoice_info_xl)


# Extract text from pdf using PyPDF2 library. Does not extract properly. Returns empty string
def read_invoice_PyPDF2(pdfFile):
    pdfFileObj = open(os.path.join(invoice_path + "\\" + pdfFile), "rb")
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
    pageObj = pdfReader.getPage(0)

    print("extracting text from " + invoice_path + "\\" + pdfFile)
    print(pageObj.extractText())


# Extract text from pdf using pdfminer3k library. Complicated but does the job
def read_invoice_pdfminer3k(pdfFile):
    fp = open(os.path.join(invoice_path + "\\" + pdfFile), "rb")

    parser = PDFParser(fp)
    doc = PDFDocument()
    parser.set_document(doc)
    doc.set_parser(parser)

    doc.initialize("")
    rsrcmgr = PDFResourceManager()
    laparams = LAParams()

    device = PDFPageAggregator(rsrcmgr, laparams=laparams)
    interpreter = PDFPageInterpreter(rsrcmgr, device)

    # Process each page contained in the document.
    invoice_text = ""
    for page in doc.get_pages():
        interpreter.process_page(page)
        layout = device.get_result()
        for lt_obj in layout:
            if isinstance(lt_obj, LTTextBox) or isinstance(lt_obj, LTTextLine):
                invoice_text += lt_obj.get_text()

    # Extract client info from the string extracted from pdf
    client = extract_info(invoice_text, client_start, client_end)
    print("client :" + client)

    # Extract invoice no from the pdf file name
    invoice_no = extract_info(str(pdfFile), invoice_start, invoice_end)
    print("invoice no :" + invoice_no)

    # Pass the client info and invoice no to the method which writes to excel file
    write_excel(client, invoice_no)


def main():
    invoice_lst = list()

    for (dirpath, dirnames,filenames) in os.walk(invoice_path):
        for filename in filenames:
            invoice_lst.append(filename)

    #print(invoice_lst)

    for index in range(len(invoice_lst)):
        print("File # {0}. Mining from the invoice file name {1}".format(index, invoice_lst[index]))
        read_invoice_pdfminer3k(invoice_lst[index])


if __name__ == "__main__":
    main()
